import openpyxl
from selenium import webdriver 
from selenium.webdriver.common.by import By
from time import sleep

Planilha_Clientes = openpyxl.load_workbook('dados_clientes.xlsx')
pagina_clientes = Planilha_Clientes['Sheet1']


for linha in pagina_clientes.iter_rows(min_row=2,values_only=True):
   Nome, Valor,CPF,Vencimento = linha

   driver = webdriver.Chrome()
   driver.get('https://consultcpf-devaprender.netlify.app/')
   sleep(5)
   campo_pesquisa = driver.find_element(By.XPATH,"//input[@id='cpfInput']")
   sleep(1)
   campo_pesquisa.clear()
   campo_pesquisa.send_keys(CPF)
   sleep(1)
   botao_pesquisar = driver.find_element(By.XPATH,"//button[@class='btn btn-custom btn-lg btn-block mt-3']")
   sleep(1)
   botao_pesquisar.click()
   sleep(4)

   status = driver.find_element(By.XPATH,"//span[@id='statusLabel']")
   if status.text == 'em dia':
     data_pagamento = driver.find_element(By.XPATH,"//p[@id='paymentDate']")
     metodo_pagamento = driver.find_element(By.XPATH,"//p[@id='paymentMethod']")
     
     data_pagamento_limpo = data_pagamento.text.split()[3]
     metodo_pagamento_limpo = metodo_pagamento.text.split()[3]
      
     planilha_fechamento = openpyxl.load_workbook('planilha_fechamento.xlsx')
     pagina_fechamento = planilha_fechamento['Sheet1']
      
     pagina_fechamento.append([Nome,Valor,CPF,Vencimento,'em dia',data_pagamento_limpo,metodo_pagamento_limpo])
     planilha_fechamento.save('planilha_fechamento.xlsx')
     
   else:
      planilha_fechamento = openpyxl.load_workbook('planilha_fechamento.xlsx')
      pagina_fechamento = planilha_fechamento['Sheet1']

      pagina_fechamento.append([Nome,Valor,CPF,Vencimento,'Pendente'])
      planilha_fechamento.save('planilha_fechamento.xlsx')
